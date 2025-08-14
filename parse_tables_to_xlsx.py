# -*- coding: utf-8 -*-
"""
テキストに含まれる独自タグから表を検出し、行/列の結合を再現して Excel (.xlsx) に出力します。
- <"図表ネーム"> …… 直近の表のシート名候補（「表」を含む場合のみ採用）
- <"表...">        …… 表の開始
- <"行"> / <"行_罫なし"> …… 1行の区切り
- <"G...">内容     …… セル。 '＝C◯_C◯' / '＝C◯_T◯' 等の数値で rowspan/colspan を表現
  例: ＝C2_C1 → 縦2行 × 横1列、 ＝C1_C2 → 縦1行 × 横2列
- <KG> は ' / ' に変換、その他の <...> は除去（<sub> 等の装飾を落とす）

使い方:
    python parse_tables_to_xlsx.py input.txt output.xlsx
"""

import sys
import re
from typing import List, Tuple, Dict, Any
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# ---------- パーサ設定 ----------
CAPTION_RE      = re.compile(r'^<"図表ネーム">(.*)$')
TABLE_START_RE  = re.compile(r'^<"表[^"]*">$')
ROW_RE          = re.compile(r'^<"行[^"]*">$')   # <"行"> と <"行_罫なし"> をまとめて扱う
CELL_RE         = re.compile(r'^<"G[^"]*">(.*)$')
SPAN_RE         = re.compile(r'＝([CT])(\d+)_([CT])(\d+)')  # ＝C2_C1 / ＝C1_T2 などから数値抽出

# ---------- ユーティリティ ----------
def strip_inline_tags(text: str) -> str:
    """<KG> を ' / ' に、その他 <...> は除去。空白を整形。"""
    text = text.replace('<KG>', ' / ')
    text = re.sub(r'<[^>]+>', '', text)
    return re.sub(r'\s+', ' ', text).strip()

def parse_span(tag: str) -> Tuple[int, int]:
    """タグ中の '＝C◯_C◯' / '＝C◯_T◯' から (rowspan, colspan) を得る。なければ (1,1)。"""
    m = SPAN_RE.search(tag)
    if not m:
        return 1, 1
    # 文字（C/T）は無視し、数値だけ使用
    rowspan = int(m.group(2))
    colspan = int(m.group(4))
    return rowspan, colspan

def parse_cell_line(line: str) -> Tuple[str, int, int]:
    """セル行から (value, rowspan, colspan) を抽出。"""
    # 先に '">…' より左側をタグとみなす
    tag = line.split('">', 1)[0]
    rs, cs = parse_span(tag)
    cm = CELL_RE.match(line)
    content = strip_inline_tags(cm.group(1)) if cm else ''
    return content, rs, cs

def sanitize_sheet_name(name: str) -> str:
    """Excelのシート名（31文字・禁止文字）に合わせて整形。"""
    name = name.replace(':', '：')
    name = re.sub(r'[\\/*?\[\]]', ' ', name)
    return name[:31] if len(name) > 31 else name

# ---------- ヘッダー深度推定 ----------
def infer_header_depth_from_first_row(rows_spec: List[List[Tuple[str, int, int]]]) -> int:
    """
    新しいヘッダー深度推定ロジック：
    表の最初の行（最初に出現する <"行"> ブロック）に含まれるセルのrowspanの最大値を返す。
    
    規則：
    - 表の最初の行に出現する各セルの "前半のC/Tに付く数値"（= 行方向の連結数 / rowspan）の最大値を header depth とする
    - 例：最初の行に ＝C2_C1, ＝C1_C2, ＝C2_C1 があれば max(2,1,2)=2 行がヘッダー
    - セルがない/数値が取れない場合は 1 を返す（フォールバック）
    - ヘッダーの次の行のA列が「分類」の場合、header depthを+1する
    
    Args:
        rows_spec: [[(val, rowspan, colspan), ...], ...] 形式のリスト
    
    Returns:
        int: ヘッダー深度（最低1）
    """
    if not rows_spec:
        return 1
    
    # 最初の行を取得
    first_row = rows_spec[0] if rows_spec else []
    
    max_rs = 1
    for cell in first_row:
        # cell は (value, rowspan, colspan) 形式
        # rowspan は2番目の要素（インデックス1）
        try:
            rs = int(cell[1]) if len(cell) > 1 else 1
            if rs > max_rs:
                max_rs = rs
        except (ValueError, TypeError, IndexError):
            # 数値に変換できない場合はスキップ
            pass
    
    # ヘッダーの次の行のA列が「分類」の場合、header depthを+1する
    if len(rows_spec) > max_rs:
        next_row = rows_spec[max_rs]
        if next_row and len(next_row) > 0:
            # A列（0番目のカラム）の値を取得
            first_cell_value = next_row[0][0] if len(next_row[0]) > 0 else ""
            # スペースを除去してから判定
            if first_cell_value and "分類" in first_cell_value.replace(" ", "").replace("　", ""):
                max_rs += 1
    
    return max_rs

# ---------- コア配置ロジック ----------
def place_cells(rows_spec: List[List[Tuple[str, int, int]]]) -> Tuple[List[List[str]], List[Tuple[int, int, int, int]]]:
    """
    ★FIX: 空文字 '' を空き判定に使わず、占有マップ(occ)で上段スパンの占有を管理。
          これにより 2段目ヘッダー（例: C / TG）が結合セルに飲み込まれず正位置に入る。
    """
    grid: List[List[str]] = []
    merges: List[Tuple[int, int, int, int]] = []
    occ: Dict[Tuple[int, int], bool] = {}  # 1-based (row, col)

    row_idx = 0
    for r_cells in rows_spec:
        row_idx += 1
        while len(grid) < row_idx:
            grid.append([])

        col = 1
        for val, rs, cs in r_cells:
            # 次の空き列を occ に基づいて探索（上段スパンで覆われたセルはスキップ）
            while occ.get((row_idx, col), False):
                col += 1
            # 表示用グリッドの拡張
            for rr in range(row_idx, row_idx + rs):
                while len(grid) < rr:
                    grid.append([])
                while len(grid[rr - 1]) < col + cs - 1:
                    grid[rr - 1].append('')
            while len(grid[row_idx - 1]) < col:
                grid[row_idx - 1].append('')
            # 値配置
            grid[row_idx - 1][col - 1] = val
            # スパン領域の占有マーキング
            for rr in range(row_idx, row_idx + rs):
                for cc in range(col, col + cs):
                    occ[(rr, cc)] = True
            if rs > 1 or cs > 1:
                merges.append((row_idx, col, row_idx + rs - 1, col + cs - 1))
            col += cs

    return grid, merges

# ---------- メインパース ----------
def parse_text_to_tables(text: str) -> List[Dict[str, Any]]:
    """
    テキスト全体を走査し、表ごとの rows_spec を構築。
    戻り値: [{'id': table_id, 'name': sheet_name, 'rows': [ [(val,rs,cs), ...], ... ]}, ...]
    """
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    tables: List[Dict[str, Any]] = []
    current_caption = None
    current_table = None
    current_row_cells: List[Tuple[str, int, int]] = []

    def flush_row():
        nonlocal current_row_cells, current_table
        if current_table is not None and current_row_cells:
            current_table['rows'].append(current_row_cells)
            current_row_cells = []

    def flush_table():
        nonlocal current_table, current_row_cells, tables
        if current_table is not None:
            flush_row()
            tables.append(current_table)
            current_table = None
            current_row_cells = []

    for ln in lines:
        mcap = CAPTION_RE.match(ln)
        if mcap:
            current_caption = strip_inline_tags(mcap.group(1))
            continue

        if TABLE_START_RE.match(ln):
            flush_table()
            table_id = ln[2:-2]  # <" と "> を剥がす
            # シート名は直近の「図表ネーム」に「表」が含まれていればそちら、無ければID
            sheet_name = current_caption if (current_caption and '表' in current_caption) else table_id
            current_table = {'id': table_id, 'name': sanitize_sheet_name(sheet_name), 'rows': []}
            continue

        if ROW_RE.match(ln):
            flush_row()
            continue

        if CELL_RE.match(ln) and current_table is not None:
            val, rs, cs = parse_cell_line(ln)
            current_row_cells.append((val, rs, cs))
            continue

        # ここに来るのは図の説明や注記など。今回は無視。

    flush_table()
    return tables

# ---------- Excel出力 ----------
def write_tables_to_xlsx(tables: List[Dict[str, Any]], out_path: str):
    wb = Workbook()
    # 既定シート削除
    wb.remove(wb.active)

    used_names = set()
    for tbl in tables:
        grid, merges = place_cells(tbl['rows'])

        # シート名重複回避
        name = tbl['name']
        base = name[:28]
        sheet_name = name
        i = 2
        while sheet_name in used_names:
            sheet_name = f"{base}_{i}"
            i += 1
        used_names.add(sheet_name)

        ws = wb.create_sheet(title=sheet_name)

        # 書き出し
        for r_idx, row in enumerate(grid, start=1):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)

        # セル結合
        for r1, c1, r2, c2 in merges:
            if not (r1 == r2 and c1 == c2):
                ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

        # 列幅をざっくり調整（全角混在は厳密でないが視認性を上げる）
        col_count = max((len(r) for r in grid), default=0)
        for c in range(1, col_count + 1):
            # その列の最大文字数
            max_len = 0
            for r in range(1, len(grid) + 1):
                v = ws.cell(row=r, column=c).value
                if isinstance(v, str):
                    max_len = max(max_len, len(v))
            # 上限を設けつつ少し余白
            width = min(50, max(8, max_len + 2))
            ws.column_dimensions[get_column_letter(c)].width = width

    wb.save(out_path)

# ---------- CLI ----------
def main():
    import argparse
    
    # CLIパーサーの設定
    parser = argparse.ArgumentParser(description='独自タグから表を検出し、行/列の結合を再現してExcel出力')
    parser.add_argument('input_file', help='入力テキストファイル')
    parser.add_argument('output_file', nargs='?', default='parsed_tables.xlsx', 
                        help='出力Excelファイル（デフォルト: parsed_tables.xlsx）')
    parser.add_argument('--header-depth', type=int, metavar='N',
                        help='ヘッダー深度を手動で指定（省略時は自動推定）')
    
    args = parser.parse_args()
    
    # header-depthのバリデーション
    manual_header_depth = None
    if args.header_depth is not None:
        if args.header_depth < 1:
            print(f"警告: --header-depth {args.header_depth} は無効です。1に矯正します。")
            manual_header_depth = 1
        else:
            manual_header_depth = args.header_depth
            print(f"手動指定: header_depth = {manual_header_depth}")

    with open(args.input_file, "r", encoding="utf-8") as f:
        text = f.read()

    tables = parse_text_to_tables(text)
    if not tables:
        print("No tables detected.")
        sys.exit(0)
    
    # ヘッダー深度情報を各テーブルに追加（ログ出力用）
    for tbl in tables:
        if manual_header_depth is not None:
            header_depth = manual_header_depth
        else:
            header_depth = infer_header_depth_from_first_row(tbl['rows'])
        print(f"Table '{tbl['name']}': header_depth = {header_depth}")

    write_tables_to_xlsx(tables, args.output_file)
    print(f"Done. Wrote {len(tables)} table(s) to: {args.output_file}")

if __name__ == "__main__":
    main()
